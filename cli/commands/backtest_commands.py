"""
Lệnh backtesting hệ thống giao dịch.
File này cung cấp giao diện dòng lệnh và các tiện ích để thực hiện backtesting,
phân tích kết quả, tối ưu hóa tham số và đánh giá chiến lược giao dịch.
"""

import os
import sys
import argparse
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import json
import logging
from typing import Dict, List, Tuple, Union, Any, Optional
from pathlib import Path
import yaml
from datetime import datetime
import concurrent.futures
from functools import partial

# Import các module từ hệ thống
from config.logging_config import get_logger
from config.constants import BacktestMetric, PositionSide, OrderType, OrderStatus
from config.system_config import get_system_config, DATA_DIR, BACKTEST_DIR

# Import các module backtesting
from backtesting.backtester import Backtester
from backtesting.strategy_tester import StrategyTester
from backtesting.performance_metrics import PerformanceMetrics
from backtesting.historical_simulator import HistoricalSimulator
from backtesting.evaluation.performance_evaluator import PerformanceEvaluator

# Thiết lập logger
logger = get_logger("backtest_command")

class BacktestCommands:
    """
    Lớp cung cấp các lệnh giao diện dòng lệnh để thực hiện backtesting.
    Cho phép người dùng chạy các backtesting, đánh giá và tối ưu hóa chiến lược.
    """
    
    def __init__(
        self,
        data_dir: Optional[Path] = None,
        output_dir: Optional[Path] = None,
        config_path: Optional[Path] = None,
        verbose: bool = True
    ):
        """
        Khởi tạo đối tượng BacktestCommands.
        
        Args:
            data_dir: Thư mục dữ liệu đầu vào
            output_dir: Thư mục đầu ra kết quả
            config_path: Đường dẫn file cấu hình
            verbose: In thông tin chi tiết
        """
        # Thiết lập logger
        self.logger = logger
        self.verbose = verbose
        
        # Thiết lập thư mục dữ liệu
        if data_dir is None:
            self.data_dir = DATA_DIR
        else:
            self.data_dir = Path(data_dir)
        
        # Thiết lập thư mục đầu ra
        if output_dir is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_dir = BACKTEST_DIR / f"backtest_{timestamp}"
        else:
            self.output_dir = Path(output_dir)
        
        # Tạo thư mục đầu ra nếu chưa tồn tại
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Thiết lập cấu hình
        self.config = get_system_config()
        
        # Tải file cấu hình nếu có
        if config_path:
            self._load_config(config_path)
        
        # Khởi tạo Backtester
        self.backtester = Backtester(
            data_dir=self.data_dir,
            output_dir=self.output_dir,
            logger=self.logger
        )
        
        # Khởi tạo PerformanceEvaluator
        self.evaluator = PerformanceEvaluator(
            results_dir=self.output_dir,
            logger=self.logger
        )
        
        # Lưu trữ chiến lược đã đăng ký
        self.strategies = {}
        
        # Lưu trữ kết quả backtesting
        self.results = {}
        
        self.logger.info(f"Đã khởi tạo BacktestCommands với output_dir={self.output_dir}")
    
    def _load_config(self, config_path: Union[str, Path]) -> None:
        """
        Tải cấu hình từ file.
        
        Args:
            config_path: Đường dẫn file cấu hình
        """
        config_path = Path(config_path)
        
        if not config_path.exists():
            self.logger.error(f"File cấu hình không tồn tại: {config_path}")
            return
        
        try:
            # Đọc file cấu hình dựa trên định dạng
            if config_path.suffix == ".json":
                with open(config_path, "r", encoding="utf-8") as f:
                    config = json.load(f)
            elif config_path.suffix in [".yaml", ".yml"]:
                with open(config_path, "r", encoding="utf-8") as f:
                    config = yaml.safe_load(f)
            else:
                self.logger.error(f"Định dạng file cấu hình không được hỗ trợ: {config_path.suffix}")
                return
            
            # Cập nhật cấu hình
            if isinstance(config, dict):
                self.config.update(config)
                self.logger.info(f"Đã tải cấu hình từ {config_path}")
            else:
                self.logger.error(f"File cấu hình không hợp lệ: {config_path}")
        
        except Exception as e:
            self.logger.error(f"Lỗi khi tải file cấu hình: {e}")
    
    def register_strategy(
        self,
        strategy_func: callable,
        strategy_name: str,
        strategy_params: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        Đăng ký chiến lược giao dịch.
        
        Args:
            strategy_func: Hàm chiến lược giao dịch
            strategy_name: Tên chiến lược
            strategy_params: Tham số chiến lược
        """
        # Đăng ký chiến lược vào Backtester
        self.backtester.register_strategy(
            strategy_func=strategy_func,
            strategy_name=strategy_name,
            strategy_params=strategy_params
        )
        
        # Lưu thông tin chiến lược
        self.strategies[strategy_name] = {
            "function": strategy_func,
            "params": strategy_params or {},
            "registered_at": datetime.now().isoformat()
        }
        
        self.logger.info(f"Đã đăng ký chiến lược '{strategy_name}'")
    
    def load_data(
        self,
        symbols: Union[str, List[str]],
        timeframe: str,
        start_date: Optional[Union[str, datetime]] = None,
        end_date: Optional[Union[str, datetime]] = None,
        file_paths: Optional[Dict[str, Path]] = None,
        data_format: str = "parquet",
        clean_data: bool = True,
        generate_features: bool = True
    ) -> Dict[str, pd.DataFrame]:
        """
        Tải dữ liệu cho backtesting.
        
        Args:
            symbols: Danh sách ký hiệu tài sản
            timeframe: Khung thời gian dữ liệu
            start_date: Ngày bắt đầu
            end_date: Ngày kết thúc
            file_paths: Dict ánh xạ symbol -> đường dẫn file
            data_format: Định dạng dữ liệu
            clean_data: Làm sạch dữ liệu
            generate_features: Tạo đặc trưng
            
        Returns:
            Dict ánh xạ symbol -> DataFrame dữ liệu
        """
        return self.backtester.load_data(
            symbols=symbols,
            timeframe=timeframe,
            start_date=start_date,
            end_date=end_date,
            file_paths=file_paths,
            data_format=data_format,
            clean_data=clean_data,
            generate_features=generate_features
        )
    
    def run_backtest(
        self,
        strategy_name: str,
        data: Union[pd.DataFrame, Dict[str, pd.DataFrame]],
        strategy_params: Optional[Dict[str, Any]] = None,
        backtest_params: Optional[Dict[str, Any]] = None,
        save_results: bool = True,
        plot_results: bool = True,
        verbose: Optional[bool] = None
    ) -> Dict[str, Any]:
        """
        Chạy backtest cho một chiến lược.
        
        Args:
            strategy_name: Tên chiến lược
            data: DataFrame hoặc Dict chứa dữ liệu
            strategy_params: Tham số chiến lược
            backtest_params: Tham số backtest
            save_results: Lưu kết quả
            plot_results: Vẽ biểu đồ kết quả
            verbose: In thông tin chi tiết
            
        Returns:
            Dict kết quả backtest
        """
        if verbose is None:
            verbose = self.verbose
        
        # Kiểm tra chiến lược
        if strategy_name not in self.strategies and strategy_name not in self.backtester.strategies:
            self.logger.error(f"Chiến lược '{strategy_name}' chưa được đăng ký")
            return {"status": "error", "message": f"Chiến lược '{strategy_name}' chưa được đăng ký"}
        
        # Chạy backtest
        result = self.backtester.run_backtest(
            strategy_name=strategy_name,
            data=data,
            strategy_params=strategy_params,
            backtest_params=backtest_params,
            verbose=verbose
        )
        
        # Lưu kết quả
        if save_results:
            self.results[strategy_name] = result
        
        # Vẽ biểu đồ kết quả
        if plot_results:
            self.plot_backtest_results(result)
        
        return result
    
    def run_multiple_backtests(
        self,
        strategies: Union[str, List[str]],
        data: Union[pd.DataFrame, Dict[str, pd.DataFrame]],
        strategy_params: Optional[Dict[str, Dict[str, Any]]] = None,
        backtest_params: Optional[Dict[str, Any]] = None,
        parallel: bool = True,
        max_workers: Optional[int] = None,
        save_results: bool = True,
        plot_comparison: bool = True,
        verbose: Optional[bool] = None
    ) -> Dict[str, Dict[str, Any]]:
        """
        Chạy backtest cho nhiều chiến lược.
        
        Args:
            strategies: Tên chiến lược hoặc danh sách tên chiến lược
            data: DataFrame hoặc Dict chứa dữ liệu
            strategy_params: Dict ánh xạ tên chiến lược -> tham số
            backtest_params: Tham số backtest chung
            parallel: Chạy song song hay tuần tự
            max_workers: Số luồng tối đa
            save_results: Lưu kết quả
            plot_comparison: Vẽ biểu đồ so sánh
            verbose: In thông tin chi tiết
            
        Returns:
            Dict ánh xạ tên chiến lược -> kết quả backtest
        """
        if verbose is None:
            verbose = self.verbose
        
        # Chạy nhiều backtests
        results = self.backtester.run_multiple_backtests(
            strategies=strategies,
            data=data,
            strategy_params=strategy_params,
            backtest_params=backtest_params,
            parallel=parallel,
            max_workers=max_workers,
            verbose=verbose
        )
        
        # Lưu kết quả
        if save_results:
            self.results.update(results)
        
        # Vẽ biểu đồ so sánh
        if plot_comparison and len(results) > 1:
            self.plot_strategy_comparison(results)
        
        return results
    
    def optimize_strategy(
        self,
        strategy_name: str,
        data: Union[pd.DataFrame, Dict[str, pd.DataFrame]],
        param_grid: Dict[str, List[Any]],
        metric_to_optimize: str = BacktestMetric.SHARPE_RATIO.value,
        max_workers: Optional[int] = None,
        use_walk_forward: bool = False,
        verbose: Optional[bool] = None
    ) -> Dict[str, Any]:
        """
        Tối ưu hóa tham số chiến lược.
        
        Args:
            strategy_name: Tên chiến lược
            data: DataFrame hoặc Dict chứa dữ liệu
            param_grid: Grid tham số cần tối ưu hóa
            metric_to_optimize: Chỉ số cần tối ưu hóa
            max_workers: Số luồng tối đa
            use_walk_forward: Sử dụng phương pháp walk-forward
            verbose: In thông tin chi tiết
            
        Returns:
            Dict kết quả tối ưu hóa
        """
        if verbose is None:
            verbose = self.verbose
        
        # Tối ưu hóa tham số
        result = self.backtester.optimize_strategy(
            strategy_name=strategy_name,
            data=data,
            param_grid=param_grid,
            metric_to_optimize=metric_to_optimize,
            max_workers=max_workers,
            use_walk_forward=use_walk_forward,
            verbose=verbose
        )
        
        # Vẽ biểu đồ kết quả tối ưu hóa
        if verbose and result["status"] == "success":
            self.plot_optimization_results(result)
        
        return result
    
    def evaluate_strategy(
        self,
        result: Optional[Dict[str, Any]] = None,
        strategy_name: Optional[str] = None,
        detailed: bool = True
    ) -> Dict[str, Any]:
        """
        Đánh giá hiệu suất của một chiến lược.
        
        Args:
            result: Kết quả backtest
            strategy_name: Tên chiến lược
            detailed: Đánh giá chi tiết hay không
            
        Returns:
            Dict chứa kết quả đánh giá
        """
        # Lấy kết quả nếu chỉ cung cấp tên chiến lược
        if result is None and strategy_name is not None:
            if strategy_name in self.results:
                result = self.results[strategy_name]
            else:
                self.logger.error(f"Không tìm thấy kết quả cho chiến lược '{strategy_name}'")
                return {"status": "error", "message": f"Không tìm thấy kết quả cho chiến lược '{strategy_name}'"}
        
        if result is None:
            self.logger.error("Cần cung cấp kết quả hoặc tên chiến lược để đánh giá")
            return {"status": "error", "message": "Cần cung cấp kết quả hoặc tên chiến lược để đánh giá"}
        
        # Đánh giá chiến lược
        evaluation = self.evaluator.evaluate_strategy(result, strategy_name, detailed)
        
        # In kết quả đánh giá
        if self.verbose:
            self.print_evaluation_summary(evaluation)
        
        return evaluation
    
    def compare_strategies(
        self,
        evaluations: Optional[Dict[str, Dict[str, Any]]] = None,
        metrics: Optional[List[str]] = None,
        sort_by: str = BacktestMetric.SHARPE_RATIO.value,
        ascending: bool = False
    ) -> pd.DataFrame:
        """
        So sánh các chiến lược dựa trên các chỉ số đã đánh giá.
        
        Args:
            evaluations: Dict ánh xạ tên chiến lược -> kết quả đánh giá
            metrics: Danh sách chỉ số cần so sánh
            sort_by: Chỉ số để sắp xếp
            ascending: Sắp xếp tăng dần hay không
            
        Returns:
            DataFrame so sánh các chiến lược
        """
        # Đánh giá tất cả chiến lược nếu chưa cung cấp
        if evaluations is None:
            # Đánh giá từng chiến lược
            evaluations = {}
            for strategy_name, result in self.results.items():
                evaluation = self.evaluate_strategy(result, strategy_name, detailed=False)
                evaluations[strategy_name] = evaluation
        
        # So sánh các chiến lược
        comparison_df = self.evaluator.compare_strategies(evaluations, metrics, sort_by, ascending)
        
        # In bảng so sánh
        if self.verbose and not comparison_df.empty:
            print("\n===== So sánh chiến lược =====")
            print(comparison_df.to_string(index=False))
            print("=" * 40)
        
        return comparison_df
    
    def generate_strategy_ranking(
        self,
        evaluations: Optional[Dict[str, Dict[str, Any]]] = None,
        metrics: Optional[List[str]] = None,
        weights: Optional[Dict[str, float]] = None,
        ascending: bool = False
    ) -> pd.DataFrame:
        """
        Tạo bảng xếp hạng các chiến lược dựa trên nhiều chỉ số.
        
        Args:
            evaluations: Dict ánh xạ tên chiến lược -> kết quả đánh giá
            metrics: Danh sách chỉ số để xếp hạng
            weights: Trọng số cho từng chỉ số
            ascending: Sắp xếp tăng dần hay không
            
        Returns:
            DataFrame bảng xếp hạng
        """
        # Đánh giá tất cả chiến lược nếu chưa cung cấp
        if evaluations is None:
            # Đánh giá từng chiến lược
            evaluations = {}
            for strategy_name, result in self.results.items():
                evaluation = self.evaluate_strategy(result, strategy_name, detailed=False)
                evaluations[strategy_name] = evaluation
        
        # Tạo bảng xếp hạng
        ranking_df = self.evaluator.generate_strategies_ranking(evaluations, metrics, weights, ascending)
        
        # In bảng xếp hạng
        if self.verbose and not ranking_df.empty:
            print("\n===== Xếp hạng chiến lược =====")
            print(ranking_df.to_string(index=False))
            print("=" * 40)
        
        return ranking_df
    
    def generate_report(
        self,
        strategy_name: str,
        output_path: Optional[Union[str, Path]] = None,
        include_plots: bool = True
    ) -> Dict[str, Any]:
        """
        Tạo báo cáo đánh giá cho một chiến lược.
        
        Args:
            strategy_name: Tên chiến lược
            output_path: Đường dẫn để lưu báo cáo
            include_plots: Bao gồm biểu đồ hay không
            
        Returns:
            Dict chứa báo cáo
        """
        # Đánh giá chiến lược nếu chưa được đánh giá
        if strategy_name not in self.evaluator.evaluated_results:
            if strategy_name in self.results:
                self.evaluate_strategy(self.results[strategy_name], strategy_name)
            else:
                self.logger.error(f"Không tìm thấy kết quả cho chiến lược '{strategy_name}'")
                return {"status": "error", "message": f"Không tìm thấy kết quả cho chiến lược '{strategy_name}'"}
        
        # Thiết lập đường dẫn mặc định nếu không được cung cấp
        if output_path is None:
            output_path = self.output_dir / f"{strategy_name}_report.json"
        
        # Tạo báo cáo
        report = self.evaluator.generate_report(strategy_name, output_path, include_plots)
        
        return report
    
    def monte_carlo_analysis(
        self,
        strategy_name: str,
        strategy_func: callable,
        strategy_params: Dict[str, Any],
        data: Union[pd.DataFrame, Dict[str, pd.DataFrame]],
        num_simulations: int = 30,
        randomize_func: Optional[callable] = None,
        save_results: bool = True,
        plot_results: bool = True
    ) -> Dict[str, Any]:
        """
        Thực hiện phân tích Monte Carlo.
        
        Args:
            strategy_name: Tên chiến lược
            strategy_func: Hàm chiến lược
            strategy_params: Tham số chiến lược
            data: Dữ liệu lịch sử
            num_simulations: Số lần mô phỏng
            randomize_func: Hàm ngẫu nhiên hóa
            save_results: Lưu kết quả
            plot_results: Vẽ biểu đồ kết quả
            
        Returns:
            Dict kết quả phân tích
        """
        if not isinstance(data, dict) or not data:
            self.logger.error("Cần cung cấp dữ liệu dưới dạng Dict chứa DataFrame")
            return {"status": "error", "message": "Cần cung cấp dữ liệu dưới dạng Dict chứa DataFrame"}
        
        # Lấy dữ liệu cho symbol đầu tiên
        symbol = list(data.keys())[0]
        df = data[symbol]
        
        # Khởi tạo HistoricalSimulator
        simulator = HistoricalSimulator(
            data={symbol: df},
            symbols=[symbol],
            initial_balance=self.backtester.config["backtest"]["initial_balance"],
            fee_rate=self.backtester.config["backtest"]["fee_rate"],
            slippage=self.backtester.config["backtest"]["slippage"],
            max_leverage=self.backtester.config["backtest"]["leverage"],
            logger=self.logger
        )
        
        # Thực hiện phân tích Monte Carlo
        analysis_result = simulator.monte_carlo_analysis(
            strategy_fn=strategy_func,
            strategy_params=strategy_params,
            num_simulations=num_simulations,
            randomize_fn=randomize_func
        )
        
        # Lưu kết quả
        if save_results:
            mc_dir = self.output_dir / strategy_name / "monte_carlo"
            mc_dir.mkdir(parents=True, exist_ok=True)
            
            # Lưu kết quả dạng JSON
            with open(mc_dir / "mc_analysis.json", "w", encoding="utf-8") as f:
                # Lọc các phần không thể serialize
                save_data = {
                    "num_simulations": analysis_result["num_simulations"],
                    "mean_final_balance": analysis_result["mean_final_balance"],
                    "std_final_balance": analysis_result["std_final_balance"],
                    "worst_case": analysis_result["worst_case"],
                    "best_case": analysis_result["best_case"],
                    "percentiles": analysis_result["percentiles"]
                }
                json.dump(save_data, f, indent=4, ensure_ascii=False)
        
        # Vẽ biểu đồ kết quả
        if plot_results:
            self.plot_monte_carlo_results(analysis_result, strategy_name)
        
        return analysis_result
    
    def walk_forward_analysis(
        self,
        strategy_name: str,
        strategy_func: callable,
        strategy_params: Dict[str, Any],
        data: Union[pd.DataFrame, Dict[str, pd.DataFrame]],
        window_size: int = 90,
        step_size: int = 30,
        optimization_func: Optional[callable] = None,
        save_results: bool = True,
        plot_results: bool = True
    ) -> Dict[str, Any]:
        """
        Thực hiện phân tích walk-forward.
        
        Args:
            strategy_name: Tên chiến lược
            strategy_func: Hàm chiến lược
            strategy_params: Tham số chiến lược
            data: Dữ liệu lịch sử
            window_size: Kích thước cửa sổ
            step_size: Kích thước bước
            optimization_func: Hàm tối ưu hóa
            save_results: Lưu kết quả
            plot_results: Vẽ biểu đồ kết quả
            
        Returns:
            Dict kết quả phân tích
        """
        if not isinstance(data, dict) or not data:
            self.logger.error("Cần cung cấp dữ liệu dưới dạng Dict chứa DataFrame")
            return {"status": "error", "message": "Cần cung cấp dữ liệu dưới dạng Dict chứa DataFrame"}
        
        # Lấy dữ liệu cho symbol đầu tiên
        symbol = list(data.keys())[0]
        df = data[symbol]
        
        # Khởi tạo HistoricalSimulator
        simulator = HistoricalSimulator(
            data={symbol: df},
            symbols=[symbol],
            initial_balance=self.backtester.config["backtest"]["initial_balance"],
            fee_rate=self.backtester.config["backtest"]["fee_rate"],
            slippage=self.backtester.config["backtest"]["slippage"],
            max_leverage=self.backtester.config["backtest"]["leverage"],
            logger=self.logger
        )
        
        # Thực hiện phân tích walk-forward
        analysis_result = simulator.walk_forward_analysis(
            strategy_fn=strategy_func,
            strategy_params=strategy_params,
            window_size=window_size,
            step_size=step_size,
            optimization_fn=optimization_func
        )
        
        # Lưu kết quả
        if save_results:
            wf_dir = self.output_dir / strategy_name / "walk_forward"
            wf_dir.mkdir(parents=True, exist_ok=True)
            
            # Lưu kết quả dạng JSON
            with open(wf_dir / "wf_analysis.json", "w", encoding="utf-8") as f:
                # Lọc các phần không thể serialize
                save_data = {
                    "windows": [{"in_sample": (w["in_sample"][0].isoformat(), w["in_sample"][1].isoformat()),
                              "out_sample": (w["out_sample"][0].isoformat(), w["out_sample"][1].isoformat())}
                             for w in analysis_result["windows"]],
                    "performance_metrics": analysis_result["performance_metrics"],
                    "param_stability": analysis_result["param_stability"]
                }
                json.dump(save_data, f, indent=4, ensure_ascii=False)
        
        # Vẽ biểu đồ kết quả
        if plot_results:
            self.plot_walk_forward_results(analysis_result, strategy_name)
        
        return analysis_result
    
    def plot_backtest_results(
        self,
        result: Dict[str, Any],
        figsize: Tuple[int, int] = (15, 10),
        save_path: Optional[Union[str, Path]] = None
    ) -> None:
        """
        Vẽ kết quả backtest.
        
        Args:
            result: Kết quả backtest
            figsize: Kích thước biểu đồ
            save_path: Đường dẫn để lưu biểu đồ
        """
        if result.get("status") != "success":
            self.logger.error("Không thể vẽ kết quả backtest không thành công")
            return
        
        strategy_name = result.get("strategy_name", "Unnamed Strategy")
        
        # Tạo thư mục để lưu biểu đồ
        if save_path is None:
            plots_dir = self.output_dir / strategy_name / "plots"
            plots_dir.mkdir(parents=True, exist_ok=True)
            save_path = plots_dir / "backtest_results.png"
        
        # Trích xuất dữ liệu
        symbol_results = result.get("symbol_results", {})
        
        if not symbol_results:
            self.logger.warning("Không có kết quả symbol để vẽ biểu đồ")
            return
        
        # Tính số lượng biểu đồ cần vẽ
        num_symbols = len(symbol_results)
        rows = min(num_symbols, 3)  # Tối đa 3 hàng
        cols = (num_symbols + rows - 1) // rows  # Số cột cần thiết
        
        # Tạo biểu đồ
        fig, axes = plt.subplots(rows * 2, cols, figsize=figsize)
        
        # Xử lý trường hợp chỉ có 1 biểu đồ
        if num_symbols == 1:
            axes = np.array(axes).reshape(2, 1)
        
        # Vẽ biểu đồ cho từng symbol
        for i, (symbol, symbol_result) in enumerate(symbol_results.items()):
            if symbol_result.get("status") != "success":
                continue
            
            row = i // cols
            col = i % cols
            
            # Trích xuất dữ liệu
            balance_history = symbol_result.get("balance_history", {})
            
            if not isinstance(balance_history, dict) or "timestamp" not in balance_history or "equity" not in balance_history:
                continue
            
            # Chuyển đổi thành DataFrame
            balance_df = pd.DataFrame({
                "timestamp": pd.to_datetime(balance_history["timestamp"]),
                "balance": balance_history["balance"],
                "equity": balance_history["equity"]
            }).set_index("timestamp")
            
            # Vẽ equity curve
            ax1 = axes[row * 2, col]
            balance_df[["equity", "balance"]].plot(ax=ax1, title=f"{symbol} - Equity Curve")
            ax1.set_ylabel("Giá trị")
            ax1.grid(True, alpha=0.3)
            
            # Vẽ drawdown
            ax2 = axes[row * 2 + 1, col]
            
            # Tính drawdown
            rolling_max = balance_df["equity"].cummax()
            drawdown = (balance_df["equity"] - rolling_max) / rolling_max
            
            drawdown.plot(ax=ax2, color="red", alpha=0.7, title=f"{symbol} - Drawdown")
            ax2.set_ylabel("Drawdown")
            ax2.grid(True, alpha=0.3)
            ax2.fill_between(drawdown.index, 0, drawdown, color="red", alpha=0.3)
        
        # Ẩn các biểu đồ thừa
        for i in range(num_symbols, rows * cols):
            row = i // cols
            col = i % cols
            if row * 2 < len(axes) and col < len(axes[0]):
                axes[row * 2, col].axis("off")
                axes[row * 2 + 1, col].axis("off")
        
        plt.tight_layout()
        
        # Lưu biểu đồ nếu cần
        if save_path:
            plt.savefig(save_path, dpi=150, bbox_inches="tight")
            self.logger.info(f"Đã lưu biểu đồ tại {save_path}")
        
        plt.close(fig)
    
    def plot_strategy_comparison(
        self,
        results: Optional[Dict[str, Dict[str, Any]]] = None,
        metrics: Optional[List[str]] = None,
        figsize: Tuple[int, int] = (12, 8),
        save_path: Optional[Union[str, Path]] = None
    ) -> None:
        """
        Vẽ biểu đồ so sánh các chiến lược.
        
        Args:
            results: Dict ánh xạ tên chiến lược -> kết quả backtest
            metrics: Danh sách chỉ số cần so sánh
            figsize: Kích thước biểu đồ
            save_path: Đường dẫn để lưu biểu đồ
        """
        # Sử dụng kết quả có sẵn nếu không được cung cấp
        if results is None:
            results = self.results
        
        # Tạo thư mục để lưu biểu đồ
        if save_path is None:
            plots_dir = self.output_dir / "comparisons"
            plots_dir.mkdir(parents=True, exist_ok=True)
            save_path = plots_dir / "strategy_comparison.png"
        
        # Tạo DataFrame so sánh
        comparison_df = self.backtester.compare_strategies(results, metrics)
        
        if comparison_df.empty:
            self.logger.warning("Không có dữ liệu để vẽ biểu đồ so sánh")
            return
        
        # Thiết lập metrics mặc định
        if metrics is None:
            metrics = [
                BacktestMetric.SHARPE_RATIO.value,
                BacktestMetric.SORTINO_RATIO.value,
                BacktestMetric.MAX_DRAWDOWN.value,
                BacktestMetric.WIN_RATE.value,
                "roi",
                BacktestMetric.PROFIT_FACTOR.value
            ]
            # Lọc các metrics có trong DataFrame
            metrics = [m for m in metrics if m in comparison_df.columns]
        
        # Tạo biểu đồ
        fig, axes = plt.subplots(len(metrics), 1, figsize=figsize)
        
        if len(metrics) == 1:
            axes = [axes]
        
        for i, metric in enumerate(metrics):
            if metric not in comparison_df.columns:
                continue
            
            # Tạo barplot
            comparison_df.plot.bar(x="strategy", y=metric, ax=axes[i], legend=False)
            
            # Tùy chỉnh biểu đồ
            axes[i].set_title(f"{metric.replace('_', ' ').title()}")
            axes[i].set_ylabel(metric)
            
            # Format giá trị thành phần trăm cho một số metrics
            if metric in ["roi", BacktestMetric.MAX_DRAWDOWN.value, BacktestMetric.WIN_RATE.value]:
                axes[i].yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.1%}"))
            
            # Thêm giá trị trên mỗi cột
            for j, v in enumerate(comparison_df[metric]):
                if pd.notna(v):
                    if metric in ["roi", BacktestMetric.MAX_DRAWDOWN.value, BacktestMetric.WIN_RATE.value]:
                        text = f"{v:.1%}"
                    else:
                        text = f"{v:.2f}"
                    axes[i].text(j, v, text, ha="center", va="bottom", fontsize=9)
        
        plt.tight_layout()
        
        # Lưu biểu đồ nếu cần
        if save_path:
            plt.savefig(save_path, dpi=150, bbox_inches="tight")
            self.logger.info(f"Đã lưu biểu đồ tại {save_path}")
        
        plt.close(fig)
    
    def plot_optimization_results(
        self,
        result: Dict[str, Any],
        figsize: Tuple[int, int] = (15, 10),
        save_path: Optional[Union[str, Path]] = None
    ) -> None:
        """
        Vẽ kết quả tối ưu hóa tham số.
        
        Args:
            result: Kết quả tối ưu hóa
            figsize: Kích thước biểu đồ
            save_path: Đường dẫn để lưu biểu đồ
        """
        if result.get("status") != "success":
            self.logger.error("Không thể vẽ kết quả tối ưu hóa không thành công")
            return
        
        strategy_name = result.get("strategy_name", "Unnamed Strategy")
        
        # Tạo thư mục để lưu biểu đồ
        if save_path is None:
            plots_dir = self.output_dir / strategy_name / "optimization"
            plots_dir.mkdir(parents=True, exist_ok=True)
            save_path = plots_dir / "optimization_results.png"
        
        # Trích xuất dữ liệu
        symbol_results = result.get("symbol_results", {})
        
        if not symbol_results:
            self.logger.warning("Không có kết quả symbol để vẽ biểu đồ")
            return
        
        # Tạo biểu đồ
        fig = plt.figure(figsize=figsize)
        
        # Tạo grid layout cho biểu đồ
        grid_size = len(symbol_results)
        grid_cols = min(3, grid_size)
        grid_rows = (grid_size + grid_cols - 1) // grid_cols
        
        gs = plt.GridSpec(grid_rows * 2, grid_cols, figure=fig)
        
        # Vẽ biểu đồ cho từng symbol
        for i, (symbol, symbol_result) in enumerate(symbol_results.items()):
            if symbol_result.get("status") != "success":
                continue
            
            row = i // grid_cols
            col = i % grid_cols
            
            # Lấy dữ liệu so sánh
            comparison_df = pd.DataFrame(symbol_result.get("all_results", []))
            
            if comparison_df.empty:
                continue
            
            # Lấy tham số và metric
            params = result.get("param_grid", {})
            metric_to_optimize = result.get("metric_to_optimize", BacktestMetric.SHARPE_RATIO.value)
            
            # Vẽ biểu đồ phân phối metric
            ax1 = fig.add_subplot(gs[row * 2, col])
            ax1.hist(comparison_df["best_metric_value"], bins=20, alpha=0.7)
            ax1.set_title(f"{symbol} - {metric_to_optimize} Distribution")
            ax1.set_xlabel(metric_to_optimize)
            ax1.set_ylabel("Frequency")
            ax1.axvline(symbol_result.get("best_metric_value", 0), color="red", linestyle="--")
            
            # Vẽ biểu đồ tham số tối ưu
            ax2 = fig.add_subplot(gs[row * 2 + 1, col])
            
            # Lấy tham số tốt nhất
            best_params = symbol_result.get("best_params", {})
            
            if best_params:
                # Vẽ lại với màu khác nhau cho từng tham số
                x = list(best_params.keys())
                y = [best_params[k] for k in x]
                
                ax2.bar(x, y, alpha=0.7)
                ax2.set_title(f"{symbol} - Best Parameters")
                ax2.set_xticklabels(x, rotation=45)
                
                # Thêm giá trị
                for i, (xi, yi) in enumerate(zip(x, y)):
                    ax2.text(i, yi, f"{yi}", ha="center", va="bottom")
        
        plt.tight_layout()
        
        # Lưu biểu đồ nếu cần
        if save_path:
            plt.savefig(save_path, dpi=150, bbox_inches="tight")
            self.logger.info(f"Đã lưu biểu đồ tại {save_path}")
        
        plt.close(fig)
    
    def plot_monte_carlo_results(
        self,
        results: Dict[str, Any],
        strategy_name: str,
        figsize: Tuple[int, int] = (15, 10),
        save_path: Optional[Union[str, Path]] = None
    ) -> None:
        """
        Vẽ kết quả phân tích Monte Carlo.
        
        Args:
            results: Kết quả phân tích
            strategy_name: Tên chiến lược
            figsize: Kích thước biểu đồ
            save_path: Đường dẫn để lưu biểu đồ
        """
        # Tạo thư mục để lưu biểu đồ
        if save_path is None:
            plots_dir = self.output_dir / strategy_name / "monte_carlo"
            plots_dir.mkdir(parents=True, exist_ok=True)
            save_path = plots_dir / "monte_carlo_results.png"
        
        # Trích xuất dữ liệu
        all_equity = results.get("all_equity")
        final_balances = results.get("final_balances", [])
        percentile_curves = results.get("percentile_curves", {})
        
        if all_equity is None or all_equity.empty:
            self.logger.warning("Không có dữ liệu equity để vẽ biểu đồ")
            return
        
        # Tạo biểu đồ
        fig, axes = plt.subplots(2, 2, figsize=figsize)
        
        # Vẽ tất cả đường equity
        ax1 = axes[0, 0]
        all_equity.plot(ax=ax1, legend=False, alpha=0.3, color="gray")
        ax1.set_title("Tất cả đường Equity")
        ax1.set_xlabel("Thời gian")
        ax1.set_ylabel("Giá trị")
        ax1.grid(True, alpha=0.3)
        
        # Vẽ đường phân vị
        ax2 = axes[0, 1]
        for name, curve in percentile_curves.items():
            if name == "median":
                ax2.plot(curve.index, curve, label=name, linewidth=2, color="blue")
            elif name == "5%":
                ax2.plot(curve.index, curve, label=name, linewidth=2, color="red")
            elif name == "95%":
                ax2.plot(curve.index, curve, label=name, linewidth=2, color="green")
            else:
                ax2.plot(curve.index, curve, label=name, linewidth=1.5, alpha=0.7)
        
        ax2.set_title("Đường phân vị Equity")
        ax2.set_xlabel("Thời gian")
        ax2.set_ylabel("Giá trị")
        ax2.grid(True, alpha=0.3)
        ax2.legend()
        
        # Vẽ histogram kết quả cuối cùng
        ax3 = axes[1, 0]
        ax3.hist(final_balances, bins=20, alpha=0.7)
        ax3.set_title("Phân phối giá trị cuối cùng")
        ax3.set_xlabel("Giá trị cuối cùng")
        ax3.set_ylabel("Tần suất")
        ax3.grid(True, alpha=0.3)
        
        # Vẽ boxplot kết quả cuối cùng
        ax4 = axes[1, 1]
        ax4.boxplot(final_balances)
        ax4.set_title("Box plot giá trị cuối cùng")
        ax4.set_ylabel("Giá trị")
        ax4.grid(True, alpha=0.3)
        
        # Thêm thông tin thống kê
        textstr = "\n".join([
            f"Số lần mô phỏng: {results.get('num_simulations', 0)}",
            f"Giá trị trung bình: {results.get('mean_final_balance', 0):.2f}",
            f"Độ lệch chuẩn: {results.get('std_final_balance', 0):.2f}",
            f"Trường hợp tệ nhất: {results.get('worst_case', 0):.2f}",
            f"Trường hợp tốt nhất: {results.get('best_case', 0):.2f}"
        ])
        
        ax4.text(1.05, 0.5, textstr, transform=ax4.transAxes,
                 bbox=dict(facecolor="white", alpha=0.7, boxstyle="round,pad=0.5"))
        
        plt.tight_layout()
        
        # Lưu biểu đồ nếu cần
        if save_path:
            plt.savefig(save_path, dpi=150, bbox_inches="tight")
            self.logger.info(f"Đã lưu biểu đồ tại {save_path}")
        
        plt.close(fig)
    
    def plot_walk_forward_results(
        self,
        results: Dict[str, Any],
        strategy_name: str,
        figsize: Tuple[int, int] = (15, 10),
        save_path: Optional[Union[str, Path]] = None
    ) -> None:
        """
        Vẽ kết quả phân tích walk-forward.
        
        Args:
            results: Kết quả phân tích
            strategy_name: Tên chiến lược
            figsize: Kích thước biểu đồ
            save_path: Đường dẫn để lưu biểu đồ
        """
        # Tạo thư mục để lưu biểu đồ
        if save_path is None:
            plots_dir = self.output_dir / strategy_name / "walk_forward"
            plots_dir.mkdir(parents=True, exist_ok=True)
            save_path = plots_dir / "walk_forward_results.png"
        
        # Trích xuất dữ liệu
        window_results = results.get("window_results", [])
        param_stability = results.get("param_stability", {})
        combined_equity_curve = results.get("combined_equity_curve")
        
        if not window_results:
            self.logger.warning("Không có kết quả cửa sổ để vẽ biểu đồ")
            return
        
        # Tạo biểu đồ
        fig, axes = plt.subplots(2, 2, figsize=figsize)
        
        # Vẽ equity curve tổng hợp
        ax1 = axes[0, 0]
        if combined_equity_curve is not None and not combined_equity_curve.empty:
            combined_equity_curve.plot(ax=ax1)
            ax1.set_title("Đường Equity tổng hợp")
            ax1.set_xlabel("Thời gian")
            ax1.set_ylabel("Giá trị")
            ax1.grid(True, alpha=0.3)
        else:
            ax1.text(0.5, 0.5, "Không có dữ liệu equity curve", 
                     horizontalalignment="center", verticalalignment="center")
        
        # Vẽ hiệu suất của từng cửa sổ
        ax2 = axes[0, 1]
        
        window_metrics = []
        for i, result in enumerate(window_results):
            if "window_index" in result:
                window_idx = result["window_index"]
            else:
                window_idx = i
            
            # Trích xuất metrics
            window_metrics.append({
                "window": window_idx,
                "roi": result.get("roi", 0),
                "sharpe": result.get("metrics", {}).get(BacktestMetric.SHARPE_RATIO.value, 0),
                "drawdown": result.get("metrics", {}).get(BacktestMetric.MAX_DRAWDOWN.value, 0)
            })
        
        if window_metrics:
            metrics_df = pd.DataFrame(window_metrics)
            
            # Vẽ ROI và Sharpe Ratio
            ax2.set_title("Hiệu suất theo cửa sổ")
            ax2.set_xlabel("Cửa sổ")
            ax2.set_ylabel("Giá trị")
            
            ax2.plot(metrics_df["window"], metrics_df["roi"], marker="o", label="ROI")
            ax2.plot(metrics_df["window"], metrics_df["sharpe"], marker="s", label="Sharpe")
            
            # Vẽ Drawdown trên trục thứ hai
            ax2_twin = ax2.twinx()
            ax2_twin.plot(metrics_df["window"], metrics_df["drawdown"], marker="^", color="red", label="Drawdown")
            ax2_twin.set_ylabel("Drawdown")
            
            # Kết hợp legend
            lines1, labels1 = ax2.get_legend_handles_labels()
            lines2, labels2 = ax2_twin.get_legend_handles_labels()
            ax2.legend(lines1 + lines2, labels1 + labels2, loc="best")
            
            ax2.grid(True, alpha=0.3)
        else:
            ax2.text(0.5, 0.5, "Không có dữ liệu hiệu suất", 
                     horizontalalignment="center", verticalalignment="center")
        
        # Vẽ độ ổn định tham số
        ax3 = axes[1, 0]
        
        if param_stability:
            param_names = list(param_stability.keys())
            positions = np.arange(len(param_names))
            
            means = [param_stability[param].get("mean", 0) for param in param_names]
            stds = [param_stability[param].get("std", 0) for param in param_names]
            
            # Vẽ bar chart
            ax3.bar(positions, means, yerr=stds, alpha=0.7, capsize=10)
            ax3.set_title("Độ ổn định tham số")
            ax3.set_ylabel("Giá trị trung bình")
            ax3.set_xticks(positions)
            ax3.set_xticklabels(param_names, rotation=45)
            ax3.grid(True, alpha=0.3)
        else:
            ax3.text(0.5, 0.5, "Không có dữ liệu độ ổn định tham số", 
                     horizontalalignment="center", verticalalignment="center")
        
        # Vẽ biểu đồ thông tin tổng hợp
        ax4 = axes[1, 1]
        
        metrics = results.get("performance_metrics", {})
        if metrics:
            # Chọn các metrics quan trọng
            key_metrics = {
                "annualized_return": metrics.get(BacktestMetric.ANNUALIZED_RETURN.value, 0),
                "sharpe_ratio": metrics.get(BacktestMetric.SHARPE_RATIO.value, 0),
                "max_drawdown": metrics.get(BacktestMetric.MAX_DRAWDOWN.value, 0),
                "win_rate": metrics.get(BacktestMetric.WIN_RATE.value, 0),
                "profit_factor": metrics.get(BacktestMetric.PROFIT_FACTOR.value, 0)
            }
            
            # Vẽ bar chart
            metric_names = list(key_metrics.keys())
            metric_values = list(key_metrics.values())
            
            ax4.bar(metric_names, metric_values, alpha=0.7)
            ax4.set_title("Chỉ số tổng hợp")
            ax4.set_ylabel("Giá trị")
            ax4.set_xticklabels(metric_names, rotation=45)
            
            # Thêm giá trị
            for i, value in enumerate(metric_values):
                if value < 0.01:
                    text = f"{value:.4f}"
                else:
                    text = f"{value:.2f}"
                ax4.text(i, value, text, ha="center", va="bottom")
            
            ax4.grid(True, alpha=0.3)
        else:
            ax4.text(0.5, 0.5, "Không có dữ liệu chỉ số tổng hợp", 
                     horizontalalignment="center", verticalalignment="center")
        
        plt.tight_layout()
        
        # Lưu biểu đồ nếu cần
        if save_path:
            plt.savefig(save_path, dpi=150, bbox_inches="tight")
            self.logger.info(f"Đã lưu biểu đồ tại {save_path}")
        
        plt.close(fig)
    
    def print_evaluation_summary(self, evaluation: Dict[str, Any]) -> None:
        """
        In tóm tắt kết quả đánh giá.
        
        Args:
            evaluation: Kết quả đánh giá
        """
        if evaluation.get("status") != "success":
            self.logger.error("Không thể in tóm tắt đánh giá không thành công")
            return
        
        strategy_name = evaluation.get("strategy_name", "Unnamed Strategy")
        metrics_summary = evaluation.get("metrics_summary", {})
        strengths = evaluation.get("strengths", [])
        weaknesses = evaluation.get("weaknesses", [])
        recommendations = evaluation.get("recommendations", [])
        
        print(f"\n{'=' * 50}")
        print(f"ĐÁNH GIÁ CHIẾN LƯỢC: {strategy_name}")
        print(f"{'=' * 50}")
        
        print("\nTÓM TẮT HIỆU SUẤT:")
        print(f"{'-' * 30}")
        if metrics_summary:
            # In các chỉ số quan trọng
            for metric_name, metric_value in metrics_summary.items():
                if metric_name in ["roi", BacktestMetric.ANNUALIZED_RETURN.value, BacktestMetric.MAX_DRAWDOWN.value, BacktestMetric.WIN_RATE.value]:
                    print(f"{metric_name}: {metric_value:.2%}")
                else:
                    print(f"{metric_name}: {metric_value:.4f}")
        else:
            print("Không có dữ liệu hiệu suất")
        
        print(f"\nĐIỂM MẠNH:")
        print(f"{'-' * 30}")
        if strengths:
            for i, strength in enumerate(strengths):
                print(f"{i+1}. {strength}")
        else:
            print("Không có điểm mạnh đáng chú ý")
        
        print(f"\nĐIỂM YẾU:")
        print(f"{'-' * 30}")
        if weaknesses:
            for i, weakness in enumerate(weaknesses):
                print(f"{i+1}. {weakness}")
        else:
            print("Không có điểm yếu đáng chú ý")
        
        print(f"\nĐỀ XUẤT CẢI TIẾN:")
        print(f"{'-' * 30}")
        if recommendations:
            for i, recommendation in enumerate(recommendations):
                print(f"{i+1}. {recommendation}")
        else:
            print("Không có đề xuất cải tiến")
        
        print(f"{'=' * 50}\n")
    
    def save_results_to_excel(
        self,
        file_path: Optional[Union[str, Path]] = None,
        include_metrics: bool = True,
        include_evaluations: bool = True,
        include_trades: bool = True
    ) -> None:
        """
        Lưu tất cả kết quả vào file Excel.
        
        Args:
            file_path: Đường dẫn file Excel
            include_metrics: Bao gồm các chỉ số
            include_evaluations: Bao gồm kết quả đánh giá
            include_trades: Bao gồm lịch sử giao dịch
        """
        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            self.logger.error("Cần cài đặt openpyxl để lưu kết quả vào Excel")
            return
        
        if not self.results:
            self.logger.warning("Không có kết quả nào để lưu")
            return
        
        # Thiết lập đường dẫn mặc định nếu không được cung cấp
        if file_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_path = self.output_dir / f"backtest_results_{timestamp}.xlsx"
        
        file_path = Path(file_path)
        
        # Tạo workbook mới
        wb = openpyxl.Workbook()
        
        # Xóa sheet mặc định
        wb.remove(wb.active)
        
        # So sánh các chiến lược
        comparison_df = self.compare_strategies(self.results)
        
        if not comparison_df.empty:
            # Tạo sheet so sánh
            ws_comparison = wb.create_sheet("Strategy Comparison")
            
            # Thêm dữ liệu
            for r in dataframe_to_rows(comparison_df, index=False, header=True):
                ws_comparison.append(r)
        
        # Xử lý từng chiến lược
        for strategy_name, result in self.results.items():
            # Tạo sheet cho chiến lược
            ws_strategy = wb.create_sheet(strategy_name[:31])  # Excel giới hạn tên sheet 31 ký tự
            
            # Thêm thông tin cơ bản
            ws_strategy.append(["Strategy", strategy_name])
            
            if "combined_result" in result and result["combined_result"]:
                combined = result["combined_result"]
                
                # Thêm các thông tin tổng hợp
                ws_strategy.append(["Initial Balance", combined.get("initial_balance", 0)])
                ws_strategy.append(["Final Balance", combined.get("final_balance", 0)])
                ws_strategy.append(["ROI", combined.get("roi", 0)])
                ws_strategy.append(["Symbols Count", combined.get("symbols_count", 0)])
                
                # Thêm các metrics nếu có
                if include_metrics and "metrics" in combined:
                    metrics = combined["metrics"]
                    
                    ws_strategy.append([])
                    ws_strategy.append(["Metrics"])
                    ws_strategy.append(["Metric", "Value"])
                    
                    for metric_name, metric_value in metrics.items():
                        ws_strategy.append([metric_name, metric_value])
            
            # Thêm kết quả đánh giá nếu có
            if include_evaluations and strategy_name in self.evaluator.evaluated_results:
                evaluation = self.evaluator.evaluated_results[strategy_name]
                
                ws_strategy.append([])
                ws_strategy.append(["Evaluation"])
                
                # Thêm điểm mạnh
                strengths = evaluation.get("strengths", [])
                if strengths:
                    ws_strategy.append(["Strengths"])
                    for i, strength in enumerate(strengths):
                        ws_strategy.append([i+1, strength])
                
                # Thêm điểm yếu
                weaknesses = evaluation.get("weaknesses", [])
                if weaknesses:
                    ws_strategy.append([])
                    ws_strategy.append(["Weaknesses"])
                    for i, weakness in enumerate(weaknesses):
                        ws_strategy.append([i+1, weakness])
                
                # Thêm đề xuất
                recommendations = evaluation.get("recommendations", [])
                if recommendations:
                    ws_strategy.append([])
                    ws_strategy.append(["Recommendations"])
                    for i, recommendation in enumerate(recommendations):
                        ws_strategy.append([i+1, recommendation])
            
            # Thêm lịch sử giao dịch nếu có
            if include_trades and "symbol_results" in result:
                symbol_results = result["symbol_results"]
                
                for symbol, symbol_result in symbol_results.items():
                    # Kiểm tra có lịch sử giao dịch không
                    if "trades" in symbol_result and symbol_result["trades"]:
                        # Tạo sheet cho giao dịch của symbol
                        sheet_name = f"{strategy_name[:15]}_{symbol[:15]}_trades"
                        ws_trades = wb.create_sheet(sheet_name[:31])
                        
                        # Chuyển đổi lịch sử giao dịch thành DataFrame
                        trades_df = pd.DataFrame(symbol_result["trades"])
                        
                        # Thêm dữ liệu
                        for r in dataframe_to_rows(trades_df, index=False, header=True):
                            ws_trades.append(r)
        
        # Lưu workbook
        try:
            wb.save(file_path)
            self.logger.info(f"Đã lưu kết quả vào {file_path}")
        except Exception as e:
            self.logger.error(f"Lỗi khi lưu file Excel: {e}")

def main():
    """
    Hàm chính để chạy ứng dụng từ dòng lệnh.
    """
    # Tạo parser cho các tham số dòng lệnh
    parser = argparse.ArgumentParser(description="Backtest các chiến lược giao dịch")
    
    # Thêm các tham số
    parser.add_argument("--data-dir", type=str, help="Thư mục dữ liệu đầu vào")
    parser.add_argument("--output-dir", type=str, help="Thư mục đầu ra kết quả")
    parser.add_argument("--config", type=str, help="Đường dẫn file cấu hình")
    parser.add_argument("--verbose", action="store_true", help="In thông tin chi tiết")
    
    # Thêm các lệnh con
    subparsers = parser.add_subparsers(dest="command", help="Lệnh cần thực hiện")
    
    # Lệnh load-data
    load_data_parser = subparsers.add_parser("load-data", help="Tải dữ liệu")
    load_data_parser.add_argument("--symbols", type=str, required=True, help="Danh sách ký hiệu tài sản, phân cách bằng dấu phẩy")
    load_data_parser.add_argument("--timeframe", type=str, required=True, help="Khung thời gian dữ liệu")
    load_data_parser.add_argument("--start-date", type=str, help="Ngày bắt đầu (YYYY-MM-DD)")
    load_data_parser.add_argument("--end-date", type=str, help="Ngày kết thúc (YYYY-MM-DD)")
    load_data_parser.add_argument("--format", type=str, default="parquet", choices=["parquet", "csv", "json"], help="Định dạng dữ liệu")
    
    # Lệnh run-backtest
    backtest_parser = subparsers.add_parser("run-backtest", help="Chạy backtest")
    backtest_parser.add_argument("--strategy", type=str, required=True, help="Tên chiến lược")
    backtest_parser.add_argument("--params", type=str, help="Tham số chiến lược (JSON string hoặc đường dẫn file)")
    backtest_parser.add_argument("--symbols", type=str, required=True, help="Danh sách ký hiệu tài sản, phân cách bằng dấu phẩy")
    backtest_parser.add_argument("--timeframe", type=str, required=True, help="Khung thời gian dữ liệu")
    backtest_parser.add_argument("--start-date", type=str, help="Ngày bắt đầu (YYYY-MM-DD)")
    backtest_parser.add_argument("--end-date", type=str, help="Ngày kết thúc (YYYY-MM-DD)")
    
    # Lệnh optimize
    optimize_parser = subparsers.add_parser("optimize", help="Tối ưu hóa tham số chiến lược")
    optimize_parser.add_argument("--strategy", type=str, required=True, help="Tên chiến lược")
    optimize_parser.add_argument("--symbols", type=str, required=True, help="Danh sách ký hiệu tài sản, phân cách bằng dấu phẩy")
    optimize_parser.add_argument("--timeframe", type=str, required=True, help="Khung thời gian dữ liệu")
    optimize_parser.add_argument("--param-grid", type=str, required=True, help="Grid tham số cần tối ưu hóa (JSON string hoặc đường dẫn file)")
    optimize_parser.add_argument("--metric", type=str, default=BacktestMetric.SHARPE_RATIO.value, help="Chỉ số cần tối ưu hóa")
    optimize_parser.add_argument("--walk-forward", action="store_true", help="Sử dụng phương pháp walk-forward")
    
    # Lệnh compare
    compare_parser = subparsers.add_parser("compare", help="So sánh các chiến lược")
    compare_parser.add_argument("--strategies", type=str, required=True, help="Danh sách tên chiến lược, phân cách bằng dấu phẩy")
    compare_parser.add_argument("--metrics", type=str, help="Danh sách chỉ số cần so sánh, phân cách bằng dấu phẩy")
    compare_parser.add_argument("--symbols", type=str, required=True, help="Danh sách ký hiệu tài sản, phân cách bằng dấu phẩy")
    compare_parser.add_argument("--timeframe", type=str, required=True, help="Khung thời gian dữ liệu")
    compare_parser.add_argument("--start-date", type=str, help="Ngày bắt đầu (YYYY-MM-DD)")
    compare_parser.add_argument("--end-date", type=str, help="Ngày kết thúc (YYYY-MM-DD)")
    
    # Lệnh evaluate
    evaluate_parser = subparsers.add_parser("evaluate", help="Đánh giá chiến lược")
    evaluate_parser.add_argument("--strategy", type=str, required=True, help="Tên chiến lược")
    evaluate_parser.add_argument("--detailed", action="store_true", help="Đánh giá chi tiết")
    evaluate_parser.add_argument("--report", action="store_true", help="Tạo báo cáo đánh giá")
    
    # Lệnh monte-carlo
    mc_parser = subparsers.add_parser("monte-carlo", help="Phân tích Monte Carlo")
    mc_parser.add_argument("--strategy", type=str, required=True, help="Tên chiến lược")
    mc_parser.add_argument("--symbol", type=str, required=True, help="Ký hiệu tài sản")
    mc_parser.add_argument("--timeframe", type=str, required=True, help="Khung thời gian dữ liệu")
    mc_parser.add_argument("--params", type=str, help="Tham số chiến lược (JSON string hoặc đường dẫn file)")
    mc_parser.add_argument("--simulations", type=int, default=30, help="Số lần mô phỏng")
    
    # Lệnh walk-forward
    wf_parser = subparsers.add_parser("walk-forward", help="Phân tích walk-forward")
    wf_parser.add_argument("--strategy", type=str, required=True, help="Tên chiến lược")
    wf_parser.add_argument("--symbol", type=str, required=True, help="Ký hiệu tài sản")
    wf_parser.add_argument("--timeframe", type=str, required=True, help="Khung thời gian dữ liệu")
    wf_parser.add_argument("--params", type=str, help="Tham số chiến lược (JSON string hoặc đường dẫn file)")
    wf_parser.add_argument("--window-size", type=int, default=90, help="Kích thước cửa sổ (ngày)")
    wf_parser.add_argument("--step-size", type=int, default=30, help="Kích thước bước (ngày)")
    
    # Phân tích tham số
    args = parser.parse_args()
    
    # Kiểm tra xem đã chỉ định lệnh chưa
    if args.command is None:
        parser.print_help()
        return
    
    # Tạo đối tượng BacktestCommands
    backtest_cmd = BacktestCommands(
        data_dir=args.data_dir,
        output_dir=args.output_dir,
        config_path=args.config,
        verbose=args.verbose
    )
    
    # Xử lý các lệnh
    if args.command == "load-data":
        # Chuyển đổi danh sách ký hiệu
        symbols = args.symbols.split(",")
        
        # Tải dữ liệu
        data = backtest_cmd.load_data(
            symbols=symbols,
            timeframe=args.timeframe,
            start_date=args.start_date,
            end_date=args.end_date,
            data_format=args.format
        )
        
        # In thông tin
        for symbol, df in data.items():
            print(f"Đã tải {len(df)} dòng dữ liệu cho {symbol}")
    
    elif args.command == "run-backtest":
        # Tải chiến lược từ module
        import importlib.util
        import sys
        
        # Tìm và tải module chiến lược
        strategy_module = None
        strategy_paths = ["strategies", "trading_strategies", "custom_strategies"]
        
        for path in strategy_paths:
            module_path = Path(path) / f"{args.strategy}.py"
            if module_path.exists():
                spec = importlib.util.spec_from_file_location(args.strategy, module_path)
                strategy_module = importlib.util.module_from_spec(spec)
                sys.modules[args.strategy] = strategy_module
                spec.loader.exec_module(strategy_module)
                break
        
        if strategy_module is None:
            print(f"Không tìm thấy file chiến lược {args.strategy}.py")
            return
        
        # Tìm hàm chiến lược
        strategy_func = None
        for attr_name in dir(strategy_module):
            if attr_name.startswith("strategy_") or attr_name == args.strategy:
                strategy_func = getattr(strategy_module, attr_name)
                if callable(strategy_func):
                    break
        
        if strategy_func is None:
            print(f"Không tìm thấy hàm chiến lược trong module {args.strategy}")
            return
        
        # Đọc tham số chiến lược
        strategy_params = {}
        if args.params:
            if os.path.exists(args.params):
                with open(args.params, "r", encoding="utf-8") as f:
                    if args.params.endswith(".json"):
                        strategy_params = json.load(f)
                    elif args.params.endswith((".yaml", ".yml")):
                        strategy_params = yaml.safe_load(f)
            else:
                try:
                    strategy_params = json.loads(args.params)
                except json.JSONDecodeError:
                    print(f"Không thể phân tích tham số JSON: {args.params}")
                    return
        
        # Đăng ký chiến lược
        backtest_cmd.register_strategy(
            strategy_func=strategy_func,
            strategy_name=args.strategy,
            strategy_params=strategy_params
        )
        
        # Chuyển đổi danh sách ký hiệu
        symbols = args.symbols.split(",")
        
        # Tải dữ liệu
        data = backtest_cmd.load_data(
            symbols=symbols,
            timeframe=args.timeframe,
            start_date=args.start_date,
            end_date=args.end_date
        )
        
        # Chạy backtest
        result = backtest_cmd.run_backtest(
            strategy_name=args.strategy,
            data=data,
            strategy_params=strategy_params
        )
        
        # In kết quả
        if result["status"] == "success" and "combined_result" in result:
            combined = result["combined_result"]
            print(f"\nKết quả tổng hợp:")
            print(f"ROI: {combined['roi']:.2%}")
            
            if "metrics" in combined:
                metrics = combined["metrics"]
                for metric_name, metric_value in metrics.items():
                    if metric_name in ["win_rate", "annualized_return", "max_drawdown"]:
                        print(f"{metric_name}: {metric_value:.2%}")
                    else:
                        print(f"{metric_name}: {metric_value:.4f}")
    
    elif args.command == "optimize":
        # Tải chiến lược từ module
        import importlib.util
        import sys
        
        # Tìm và tải module chiến lược
        strategy_module = None
        strategy_paths = ["strategies", "trading_strategies", "custom_strategies"]
        
        for path in strategy_paths:
            module_path = Path(path) / f"{args.strategy}.py"
            if module_path.exists():
                spec = importlib.util.spec_from_file_location(args.strategy, module_path)
                strategy_module = importlib.util.module_from_spec(spec)
                sys.modules[args.strategy] = strategy_module
                spec.loader.exec_module(strategy_module)
                break
        
        if strategy_module is None:
            print(f"Không tìm thấy file chiến lược {args.strategy}.py")
            return
        
        # Tìm hàm chiến lược
        strategy_func = None
        for attr_name in dir(strategy_module):
            if attr_name.startswith("strategy_") or attr_name == args.strategy:
                strategy_func = getattr(strategy_module, attr_name)
                if callable(strategy_func):
                    break
        
        if strategy_func is None:
            print(f"Không tìm thấy hàm chiến lược trong module {args.strategy}")
            return
        
        # Đọc param grid
        param_grid = {}
        if os.path.exists(args.param_grid):
            with open(args.param_grid, "r", encoding="utf-8") as f:
                if args.param_grid.endswith(".json"):
                    param_grid = json.load(f)
                elif args.param_grid.endswith((".yaml", ".yml")):
                    param_grid = yaml.safe_load(f)
        else:
            try:
                param_grid = json.loads(args.param_grid)
            except json.JSONDecodeError:
                print(f"Không thể phân tích param grid JSON: {args.param_grid}")
                return
        
        # Đăng ký chiến lược
        backtest_cmd.register_strategy(
            strategy_func=strategy_func,
            strategy_name=args.strategy
        )
        
        # Chuyển đổi danh sách ký hiệu
        symbols = args.symbols.split(",")
        
        # Tải dữ liệu
        data = backtest_cmd.load_data(
            symbols=symbols,
            timeframe=args.timeframe
        )
        
        # Chạy tối ưu hóa
        result = backtest_cmd.optimize_strategy(
            strategy_name=args.strategy,
            data=data,
            param_grid=param_grid,
            metric_to_optimize=args.metric,
            use_walk_forward=args.walk_forward
        )
        
        # In kết quả
        if result["status"] == "success":
            print(f"\nTham số tối ưu: {result['best_params']}")
            print(f"Giá trị {args.metric}: {result['best_metric_value']}")
    
    elif args.command == "compare":
        # Chuyển đổi danh sách chiến lược
        strategies = args.strategies.split(",")
        
        # Tìm và tải các module chiến lược
        import importlib.util
        import sys
        
        strategy_funcs = {}
        strategy_paths = ["strategies", "trading_strategies", "custom_strategies"]
        
        for strategy_name in strategies:
            strategy_module = None
            
            for path in strategy_paths:
                module_path = Path(path) / f"{strategy_name}.py"
                if module_path.exists():
                    spec = importlib.util.spec_from_file_location(strategy_name, module_path)
                    strategy_module = importlib.util.module_from_spec(spec)
                    sys.modules[strategy_name] = strategy_module
                    spec.loader.exec_module(strategy_module)
                    break
            
            if strategy_module is None:
                print(f"Không tìm thấy file chiến lược {strategy_name}.py")
                continue
            
            # Tìm hàm chiến lược
            strategy_func = None
            for attr_name in dir(strategy_module):
                if attr_name.startswith("strategy_") or attr_name == strategy_name:
                    strategy_func = getattr(strategy_module, attr_name)
                    if callable(strategy_func):
                        strategy_funcs[strategy_name] = strategy_func
                        break
            
            if strategy_name not in strategy_funcs:
                print(f"Không tìm thấy hàm chiến lược trong module {strategy_name}")
        
        if not strategy_funcs:
            print("Không tìm thấy chiến lược nào")
            return
        
        # Đăng ký các chiến lược
        for strategy_name, strategy_func in strategy_funcs.items():
            backtest_cmd.register_strategy(
                strategy_func=strategy_func,
                strategy_name=strategy_name
            )
        
        # Chuyển đổi danh sách ký hiệu và metrics
        symbols = args.symbols.split(",")
        metrics = args.metrics.split(",") if args.metrics else None
        
        # Tải dữ liệu
        data = backtest_cmd.load_data(
            symbols=symbols,
            timeframe=args.timeframe,
            start_date=args.start_date,
            end_date=args.end_date
        )
        
        # Chạy backtest cho tất cả chiến lược
        results = backtest_cmd.run_multiple_backtests(
            strategies=list(strategy_funcs.keys()),
            data=data
        )
        
        # So sánh chiến lược
        comparison_df = backtest_cmd.compare_strategies(results, metrics)
        
        # In kết quả
        if not comparison_df.empty:
            print("\nBảng so sánh chiến lược:")
            print(comparison_df.to_string(index=False))
    
    elif args.command == "evaluate":
        # Đánh giá chiến lược
        if args.strategy in backtest_cmd.results:
            evaluation = backtest_cmd.evaluate_strategy(
                strategy_name=args.strategy,
                detailed=args.detailed
            )
            
            # Tạo báo cáo nếu yêu cầu
            if args.report and evaluation["status"] == "success":
                report = backtest_cmd.generate_report(args.strategy)
                print(f"Đã tạo báo cáo đánh giá tại {report.get('report_path', 'N/A')}")
        else:
            print(f"Không tìm thấy kết quả cho chiến lược {args.strategy}")
    
    elif args.command == "monte-carlo":
        # Tải chiến lược từ module
        import importlib.util
        import sys
        
        # Tìm và tải module chiến lược
        strategy_module = None
        strategy_paths = ["strategies", "trading_strategies", "custom_strategies"]
        
        for path in strategy_paths:
            module_path = Path(path) / f"{args.strategy}.py"
            if module_path.exists():
                spec = importlib.util.spec_from_file_location(args.strategy, module_path)
                strategy_module = importlib.util.module_from_spec(spec)
                sys.modules[args.strategy] = strategy_module
                spec.loader.exec_module(strategy_module)
                break
        
        if strategy_module is None:
            print(f"Không tìm thấy file chiến lược {args.strategy}.py")
            return
        
        # Tìm hàm chiến lược
        strategy_func = None
        for attr_name in dir(strategy_module):
            if attr_name.startswith("strategy_") or attr_name == args.strategy:
                strategy_func = getattr(strategy_module, attr_name)
                if callable(strategy_func):
                    break
        
        if strategy_func is None:
            print(f"Không tìm thấy hàm chiến lược trong module {args.strategy}")
            return
        
        # Đọc tham số chiến lược
        strategy_params = {}
        if args.params:
            if os.path.exists(args.params):
                with open(args.params, "r", encoding="utf-8") as f:
                    if args.params.endswith(".json"):
                        strategy_params = json.load(f)
                    elif args.params.endswith((".yaml", ".yml")):
                        strategy_params = yaml.safe_load(f)
            else:
                try:
                    strategy_params = json.loads(args.params)
                except json.JSONDecodeError:
                    print(f"Không thể phân tích tham số JSON: {args.params}")
                    return
        
        # Tải dữ liệu
        data = backtest_cmd.load_data(
            symbols=[args.symbol],
            timeframe=args.timeframe
        )
        
        # Chạy phân tích Monte Carlo
        result = backtest_cmd.monte_carlo_analysis(
            strategy_name=args.strategy,
            strategy_func=strategy_func,
            strategy_params=strategy_params,
            data=data,
            num_simulations=args.simulations
        )
        
        # In kết quả
        print(f"\nKết quả phân tích Monte Carlo ({args.simulations} mô phỏng):")
        print(f"Giá trị trung bình: {result['mean_final_balance']:.2f}")
        print(f"Độ lệch chuẩn: {result['std_final_balance']:.2f}")
        print(f"Trường hợp tệ nhất: {result['worst_case']:.2f}")
        print(f"Trường hợp tốt nhất: {result['best_case']:.2f}")
        
        for pct, value in result["percentiles"].items():
            print(f"Phân vị {pct}: {value:.2f}")
    
    elif args.command == "walk-forward":
        # Tải chiến lược từ module
        import importlib.util
        import sys
        
        # Tìm và tải module chiến lược
        strategy_module = None
        strategy_paths = ["strategies", "trading_strategies", "custom_strategies"]
        
        for path in strategy_paths:
            module_path = Path(path) / f"{args.strategy}.py"
            if module_path.exists():
                spec = importlib.util.spec_from_file_location(args.strategy, module_path)
                strategy_module = importlib.util.module_from_spec(spec)
                sys.modules[args.strategy] = strategy_module
                spec.loader.exec_module(strategy_module)
                break
        
        if strategy_module is None:
            print(f"Không tìm thấy file chiến lược {args.strategy}.py")
            return
        
        # Tìm hàm chiến lược
        strategy_func = None
        for attr_name in dir(strategy_module):
            if attr_name.startswith("strategy_") or attr_name == args.strategy:
                strategy_func = getattr(strategy_module, attr_name)
                if callable(strategy_func):
                    break
        
        if strategy_func is None:
            print(f"Không tìm thấy hàm chiến lược trong module {args.strategy}")
            return
        
        # Đọc tham số chiến lược
        strategy_params = {}
        if args.params:
            if os.path.exists(args.params):
                with open(args.params, "r", encoding="utf-8") as f:
                    if args.params.endswith(".json"):
                        strategy_params = json.load(f)
                    elif args.params.endswith((".yaml", ".yml")):
                        strategy_params = yaml.safe_load(f)
            else:
                try:
                    strategy_params = json.loads(args.params)
                except json.JSONDecodeError:
                    print(f"Không thể phân tích tham số JSON: {args.params}")
                    return
        
        # Tải dữ liệu
        data = backtest_cmd.load_data(
            symbols=[args.symbol],
            timeframe=args.timeframe
        )
        
        # Chạy phân tích walk-forward
        result = backtest_cmd.walk_forward_analysis(
            strategy_name=args.strategy,
            strategy_func=strategy_func,
            strategy_params=strategy_params,
            data=data,
            window_size=args.window_size,
            step_size=args.step_size
        )
        
        # In kết quả
        print(f"\nKết quả phân tích walk-forward:")
        
        metrics = result.get("performance_metrics", {})
        if metrics:
            print(f"Đánh giá tổng hợp:")
            for metric_name, metric_value in metrics.items():
                if metric_name in ["win_rate", "annualized_return", "max_drawdown"]:
                    print(f"{metric_name}: {metric_value:.2%}")
                else:
                    print(f"{metric_name}: {metric_value:.4f}")
        
        param_stability = result.get("param_stability", {})
        if param_stability:
            print(f"\nĐộ ổn định tham số:")
            for param_name, param_stats in param_stability.items():
                mean = param_stats.get("mean", 0)
                std = param_stats.get("std", 0)
                print(f"{param_name}: {mean:.4f} ± {std:.4f}")

def setup_backtest_parser(subparsers):
    """
    Thiết lập parser cho các lệnh backtest.
    
    Args:
        subparsers: Đối tượng subparsers từ argparse
    """
    # Tạo parser cho nhóm lệnh backtest
    backtest_parser = subparsers.add_parser('backtest', help='Các lệnh backtest chiến lược giao dịch')
    backtest_subparsers = backtest_parser.add_subparsers(dest='backtest_command', help='Lệnh backtest cụ thể')
    
    # Parser cho lệnh run
    run_parser = backtest_subparsers.add_parser('run', help='Chạy backtest chiến lược')
    run_parser.add_argument("--strategy", "-s", type=str, required=True, help="Tên chiến lược giao dịch")
    run_parser.add_argument("--data-file", "-f", type=str, required=True, help="Đường dẫn file dữ liệu")
    run_parser.add_argument("--symbol", type=str, help="Cặp giao dịch")
    run_parser.add_argument("--timeframe", type=str, default="1h", help="Khung thời gian")
    run_parser.add_argument("--capital", type=float, default=10000.0, help="Vốn ban đầu")
    run_parser.add_argument("--commission", type=float, default=0.001, help="Phí giao dịch (0.1%)")
    run_parser.add_argument("--start-date", type=str, help="Ngày bắt đầu (YYYY-MM-DD)")
    run_parser.add_argument("--end-date", type=str, help="Ngày kết thúc (YYYY-MM-DD)")
    run_parser.add_argument("--output-dir", "-o", type=str, help="Thư mục lưu kết quả")
    run_parser.add_argument("--plot/--no-plot", dest="plot", action="store_true", default=True, help="Hiển thị biểu đồ kết quả")
    run_parser.add_argument("--verbose", "-v", action="count", default=0, help="Mức độ chi tiết của log (0-2)")
    
    # Parser cho lệnh compare
    compare_parser = backtest_subparsers.add_parser('compare', help='So sánh nhiều chiến lược')
    compare_parser.add_argument("--strategies", "-s", nargs='+', required=True, help="Danh sách chiến lược cần so sánh")
    compare_parser.add_argument("--data-file", "-f", type=str, required=True, help="Đường dẫn file dữ liệu")
    compare_parser.add_argument("--symbol", type=str, help="Cặp giao dịch")
    compare_parser.add_argument("--timeframe", type=str, default="1h", help="Khung thời gian")
    compare_parser.add_argument("--capital", type=float, default=10000.0, help="Vốn ban đầu")
    compare_parser.add_argument("--commission", type=float, default=0.001, help="Phí giao dịch (0.1%)")
    compare_parser.add_argument("--start-date", type=str, help="Ngày bắt đầu (YYYY-MM-DD)")
    compare_parser.add_argument("--end-date", type=str, help="Ngày kết thúc (YYYY-MM-DD)")
    compare_parser.add_argument("--output-dir", "-o", type=str, help="Thư mục lưu kết quả")
    compare_parser.add_argument("--plot/--no-plot", dest="plot", action="store_true", default=True, help="Hiển thị biểu đồ kết quả")
    compare_parser.add_argument("--verbose", "-v", action="count", default=0, help="Mức độ chi tiết của log (0-2)")
    
    # Parser cho lệnh optimize
    optimize_parser = backtest_subparsers.add_parser('optimize', help='Tối ưu hóa tham số chiến lược')
    optimize_parser.add_argument("--strategy", "-s", type=str, required=True, help="Tên chiến lược giao dịch")
    optimize_parser.add_argument("--data-file", "-f", type=str, required=True, help="Đường dẫn file dữ liệu")
    optimize_parser.add_argument("--symbol", type=str, help="Cặp giao dịch")
    optimize_parser.add_argument("--timeframe", type=str, default="1h", help="Khung thời gian")
    optimize_parser.add_argument("--params", "-p", type=str, required=True, help="File JSON chứa tham số và phạm vi tối ưu")
    optimize_parser.add_argument("--metric", "-m", type=str, default="sharpe_ratio", help="Chỉ số tối ưu hóa")
    optimize_parser.add_argument("--method", type=str, default="grid", choices=["grid", "random", "bayesian"], help="Phương pháp tối ưu hóa")
    optimize_parser.add_argument("--iterations", "-i", type=int, default=100, help="Số lần lặp tối ưu hóa")
    optimize_parser.add_argument("--output-dir", "-o", type=str, help="Thư mục lưu kết quả")
    optimize_parser.add_argument("--verbose", "-v", action="count", default=0, help="Mức độ chi tiết của log (0-2)")
    
    return backtest_parser

def handle_backtest_command(args):
    """
    Xử lý các lệnh liên quan đến backtest.
    
    Args:
        args: Đối tượng chứa các tham số dòng lệnh
        
    Returns:
        int: Mã trạng thái (0 nếu thành công, khác 0 nếu lỗi)
    """
    if args.backtest_command == 'run':
        from backtesting.backtester import Backtester
        
        # Tạo đối tượng backtester
        backtester = Backtester()
        
        # Thiết lập tham số
        backtester.set_strategy(args.strategy)
        backtester.set_data(args.data_file, args.symbol, args.timeframe)
        backtester.set_capital(args.capital)
        backtester.set_commission(args.commission)
        
        if args.start_date:
            backtester.set_start_date(args.start_date)
        if args.end_date:
            backtester.set_end_date(args.end_date)
        
        # Chạy backtest
        results = backtester.run()
        
        # Lưu kết quả nếu cần
        if args.output_dir:
            backtester.save_results(args.output_dir)
        
        # Hiển thị biểu đồ nếu cần
        if args.plot:
            backtester.plot_results()
        
        return 0
    
    elif args.backtest_command == 'compare':
        from backtesting.strategy_tester import StrategyTester
        
        # Tạo đối tượng strategy tester
        tester = StrategyTester()
        
        # Thiết lập tham số
        tester.set_strategies(args.strategies)
        tester.set_data(args.data_file, args.symbol, args.timeframe)
        tester.set_capital(args.capital)
        tester.set_commission(args.commission)
        
        if args.start_date:
            tester.set_start_date(args.start_date)
        if args.end_date:
            tester.set_end_date(args.end_date)
        
        # Chạy so sánh
        results = tester.compare()
        
        # Lưu kết quả nếu cần
        if args.output_dir:
            tester.save_results(args.output_dir)
        
        # Hiển thị biểu đồ nếu cần
        if args.plot:
            tester.plot_comparison()
        
        return 0
    
    elif args.backtest_command == 'optimize':
        from backtesting.strategy_optimizer import StrategyOptimizer
        
        # Tạo đối tượng optimizer
        optimizer = StrategyOptimizer()
        
        # Thiết lập tham số
        optimizer.set_strategy(args.strategy)
        optimizer.set_data(args.data_file, args.symbol, args.timeframe)
        optimizer.set_params_from_file(args.params)
        optimizer.set_metric(args.metric)
        optimizer.set_method(args.method)
        optimizer.set_iterations(args.iterations)
        
        # Chạy tối ưu hóa
        results = optimizer.optimize()
        
        # Lưu kết quả nếu cần
        if args.output_dir:
            optimizer.save_results(args.output_dir)
        
        return 0
    
    else:
        print(f"Lệnh backtest không hợp lệ: {args.backtest_command}")
        return 1

if __name__ == "__main__":
    main()